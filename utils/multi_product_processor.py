"""多产品Excel处理器 - 将多个产品Excel文件合并并重新格式化"""
import os
import pandas as pd
import glob
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from .periodic_buy_calculator import PeriodicBuyCalculator
from .buy_rules import BuyRule


class MultiProductExcelProcessor:
    """多产品Excel处理器"""

    def __init__(self, products_pattern: str, start_date: str = None, buy_rule: BuyRule = None): # type: ignore
        """
        初始化处理器

        Args:
            products_pattern: 产品文件的模式（如 "买入平均收益_净值列表/日度净值_产品*.xlsx"）
            start_date: 统一的买入起点日期，格式：YYYYMMDD（可选）
            buy_rule: 买入规则实例（可选，如果提供则进行收益计算）
        """
        self.products_pattern = products_pattern
        self.start_date = start_date
        self.buy_rule = buy_rule
        self.products_data = {}
        self.calculation_results = {}

    def load_products(self):
        """
        加载所有符合模式的产品文件

        Returns:
            dict: 产品数据字典
        """
        files = glob.glob(self.products_pattern)
        
        if not files:
            raise ValueError(f"未找到匹配模式的文件: {self.products_pattern}")
        
        print(f"找到 {len(files)} 个产品文件")
        
        for file_path in sorted(files):
            try:
                self._load_single_file(file_path)
            except Exception as e:
                print(f"  ⚠️  加载失败: {os.path.basename(file_path)} - {str(e)}")
        
        print(f"成功加载 {len(self.products_data)} 个产品\n")
        return self.products_data

    def _load_single_file(self, file_path: str):
        """
        加载单个产品文件

        Args:
            file_path: 文件路径
        """
        # 读取原始数据，不使用header
        raw_df = pd.read_excel(file_path, header=None)

        # 获取产品信息（A1产品名称，B1产品代码）
        product_name = str(raw_df.iloc[0, 0]).strip()  # A1
        product_code = str(raw_df.iloc[0, 1]).strip()  # B1

        # 获取数据部分（A2行是标题，A3行开始是数据）
        data_df = raw_df.iloc[2:].copy()
        data_df.columns = ['日期', '单位净值', '累计净值']

        # 清理空行
        data_df = data_df.dropna(subset=['日期'])

        # 转换日期格式
        data_df['日期'] = pd.to_datetime(data_df['日期'])

        # 按日期降序排列（保持原始顺序）
        data_df = data_df.sort_values('日期', ascending=False).reset_index(drop=True)

        print(f"  ✓ {product_name} ({len(data_df)} 条数据)")

        # 存储产品数据
        self.products_data[product_name] = {
            'product_name': product_name,
            'product_code': product_code,
            'data': data_df,
            'file_path': file_path
        }

    def filter_by_start_date(self):
        """根据起点日期过滤数据"""
        if not self.start_date:
            print("未指定起点日期，保留所有数据\n")
            return

        try:
            start_dt = pd.to_datetime(self.start_date, format='%Y%m%d')
            print(f"正在过滤数据，起点日期: {start_dt.strftime('%Y-%m-%d')}")

            for product_name, product_info in self.products_data.items():
                data_df = product_info['data']

                # 过滤数据（保留start_date及之后的数据）
                filtered_df = data_df[data_df['日期'] >= start_dt].copy()

                original_count = len(data_df)
                filtered_count = len(filtered_df)

                print(f"  - {product_name}: {original_count} -> {filtered_count} 条数据")

                # 更新数据
                product_info['data'] = filtered_df

            print()
        except Exception as e:
            print(f"警告: 过滤日期时出错 - {str(e)}")
            print("将保留所有数据\n")

    def save_to_excel(self, output_file: str):
        """
        保存为标准格式的Excel文件
        每个sheet使用产品简称

        Args:
            output_file: 输出文件路径
        """
        print(f"正在保存文件: {output_file}")

        # 创建新的Excel文件
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore # 删除默认的sheet

        for product_name, product_info in self.products_data.items():
            # 使用产品简称作为sheet名称（Excel sheet名称最多31字符）
            sheet_name = product_name[:31] if len(product_name) > 31 else product_name

            # 创建新sheet
            ws = wb.create_sheet(title=sheet_name)

            # 写入产品信息（A1: 产品名称, B1: 产品代码）
            ws['A1'] = product_info['product_name']
            ws['B1'] = product_info['product_code']

            # 写入列标题（A2-C2）
            ws['A2'] = '日期'
            ws['B2'] = '单位净值'
            ws['C2'] = '累计净值'

            # 获取数据并按日期升序排列（从早到晚）
            data_df = product_info['data'].sort_values('日期', ascending=True)

            # 写入数据（从A3开始）
            for idx, row in enumerate(data_df.itertuples(index=False), start=3):
                # 转换日期格式为YYYYMMDD（整数）
                date_int = int(row.日期.strftime('%Y%m%d'))
                ws[f'A{idx}'] = date_int
                ws[f'B{idx}'] = row.单位净值
                ws[f'C{idx}'] = row.累计净值

            # 调整列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12

            print(f"  ✓ 已创建sheet: {sheet_name} ({len(data_df)} 条数据)")

        # 保存文件
        wb.save(output_file)
        print(f"\n✅ 文件保存成功: {output_file}")

    def calculate_periodic_returns(self, output_file: str = None): # type: ignore
        """
        计算周期性买入收益（如果提供了买入规则）

        Args:
            output_file: 输出文件路径（可选）
        """
        if not self.buy_rule:
            print("未指定买入规则，跳过计算\n")
            return

        print("=" * 70)
        print("计算周期性买入收益")
        print("=" * 70)

        if not output_file:
            output_dir = os.path.dirname(self.products_pattern)
            output_file = os.path.join(output_dir, f"买入收益_{self.buy_rule.get_rule_name()}.xlsx")

        # 创建Excel工作簿
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for product_name, product_info in self.products_data.items():
                try:
                    # 创建临时Excel文件用于计算
                    temp_file = f"_temp_{product_name}.xlsx"
                    self._save_product_temp_file(product_name, product_info, temp_file)

                    # 使用PeriodicBuyCalculator计算
                    calculator = PeriodicBuyCalculator(temp_file, self.buy_rule)
                    results_df = calculator.calculate_buy_returns()

                    if not results_df.empty:
                        # 格式化百分比列
                        results_df_output = results_df.copy()
                        results_df_output['每日涨跌幅'] = results_df_output['每日涨跌幅'].apply(lambda x: f"{x:.2%}")
                        results_df_output['区间收益率'] = results_df_output['区间收益率'].apply(lambda x: f"{x:.2%}")
                        results_df_output['区间年化收益率'] = results_df_output['区间年化收益率'].apply(lambda x: f"{x:.2%}")

                        # 保存到sheet（产品简称作为sheet名）
                        sheet_name = product_name[:31] if len(product_name) > 31 else product_name
                        results_df_output.to_excel(writer, sheet_name=sheet_name, index=False)

                        # 调整列宽
                        ws = writer.sheets[sheet_name]
                        for i, col in enumerate(results_df_output.columns, 1):
                            max_len = max(len(str(col)), 15)
                            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

                        print(f"  ✓ {product_name}: {len(results_df)} 个买入日期")

                        # 存储计算结果供后续使用
                        self.calculation_results[product_name] = {
                            'calculator': calculator,
                            'results': results_df
                        }
                    else:
                        print(f"  ⚠️  {product_name}: 无符合条件的买入日期")

                    # 删除临时文件
                    if os.path.exists(temp_file):
                        os.remove(temp_file)

                except Exception as e:
                    print(f"  ❌ {product_name}: 计算失败 - {str(e)}")

        print(f"\n✅ 收益计算结果保存: {output_file}")

    def _save_product_temp_file(self, product_name: str, product_info: dict, temp_file: str):
        """
        保存单个产品的临时Excel文件供计算使用

        Args:
            product_name: 产品名称
            product_info: 产品数据字典
            temp_file: 临时文件路径
        """
        wb = Workbook()
        ws = wb.active

        # 写入产品信息
        ws['A1'] = product_info['product_name'] # type: ignore
        ws['B1'] = product_info['product_code'] # type: ignore

        # 写入列标题
        ws['A2'] = '日期' # type: ignore
        ws['B2'] = '单位净值' # type: ignore
        ws['C2'] = '累计净值' # type: ignore

        # 获取数据并按日期升序排列
        data_df = product_info['data'].sort_values('日期', ascending=True)

        # 写入数据
        for idx, row in enumerate(data_df.itertuples(index=False), start=3):
            date_int = int(row.日期.strftime('%Y%m%d'))
            ws[f'A{idx}'] = date_int # type: ignore
            ws[f'B{idx}'] = row.单位净值 # type: ignore
            ws[f'C{idx}'] = row.累计净值 # type: ignore

        wb.save(temp_file)

    def process(self, output_file: str = None, calculate_returns: bool = False, returns_output_file: str = None): # type: ignore
        """
        执行完整的处理流程

        Args:
            output_file: 输出文件路径（可选）
            calculate_returns: 是否计算周期性买入收益
            returns_output_file: 收益计算结果文件路径（可选）
        """
        # 1. 加载数据
        print("=" * 70)
        print("加载产品数据")
        print("=" * 70)
        self.load_products()

        # 2. 根据起点日期过滤
        print("=" * 70)
        print("数据过滤")
        print("=" * 70)
        self.filter_by_start_date()

        # 3. 确定输出文件路径
        if not output_file:
            # 默认保存为日度净值_合并.xlsx
            output_dir = os.path.dirname(self.products_pattern)
            output_file = os.path.join(output_dir, "日度净值_合并.xlsx")

        # 4. 保存合并文件
        print("=" * 70)
        print("保存合并文件")
        print("=" * 70)
        self.save_to_excel(output_file)

        # 5. 计算周期性买入收益（如果指定了规则）
        if calculate_returns and self.buy_rule:
            self.calculate_periodic_returns(returns_output_file)

        print("=" * 70)
        print("处理完成！")
        print("=" * 70)
