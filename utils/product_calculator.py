"""产品净值计算器类"""
import pandas as pd
from openpyxl.utils import get_column_letter


class ProductNetValueCalculator:
    """产品净值数据计算器（支持多产品格式）"""

    def __init__(self, file_path: str, risk_free_rate: float = 0.02):
        """
        初始化计算器

        Args:
            file_path: Excel文件路径
            risk_free_rate: 无风险利率，用于计算夏普比率
        """
        self.file_path: str = file_path
        self.risk_free_rate: float = risk_free_rate
        self.df: pd.DataFrame = pd.DataFrame()
        self.metrics: dict = {}
        self.products: dict = {}
        self._load_data()

    def _load_data(self):
        """
        加载并预处理数据
        Excel格式：
        - A1: 产品名称（如 "锐进58源乐晟尊享A"）
        - B1: 产品代码（如 "XA1796"）
        - A2-C2: 列标题（日期、单位净值、累计净值）
        - A3起: 数据
        """
        # 读取原始数据，不使用header
        raw_df = pd.read_excel(self.file_path, header=None)

        # 获取产品信息（A1产品名称，B1产品代码）
        product_name = raw_df.iloc[0, 0]  # A1
        product_code = raw_df.iloc[0, 1]  # B1

        # 获取数据部分（A2行是标题，A3行开始是数据）
        data_df = raw_df.iloc[2:].copy()
        data_df.columns = ['日期', '单位净值', '累计净值']

        # 清理空行
        data_df = data_df.dropna(subset=['日期'])

        # 转换日期
        data_df['日期'] = pd.to_datetime(data_df['日期'], format='%Y%m%d')
        data_df.set_index("日期", inplace=True)

        self.df = data_df

        # 存储产品信息
        if pd.notna(product_name) and pd.notna(product_code):
            self.products[str(product_code).strip()] = {
                'name': str(product_name).strip(),
                'code': str(product_code).strip()
            }

    def calculate_weekly_return(self):
        """计算周收益率（倒序数据：最新日期在前）"""
        # 倒序数据：iloc[0]是最新，iloc[-1]是最早
        # shift(-1) 获取下一行（更早的日期）作为"上周"净值
        self.df["上周净值"] = self.df["单位净值"].shift(-1)
        self.df["周收益率"] = self.df["单位净值"] / self.df["上周净值"] - 1
        self.df['周收益率'] = self.df['周收益率'].fillna(0)

    def calculate_all_return(self):
        """计算成立以来收益率（倒序数据：最新日期在前）"""
        # iloc[0] = 最新净值, iloc[-1] = 最早净值
        all_return = self.df['单位净值'].iloc[0] / self.df['单位净值'].iloc[-1] - 1
        self.metrics['all_return'] = all_return
        return all_return

    def calculate_annual_return(self):
        """计算年化收益率"""
        all_return = self.metrics.get('all_return', self.calculate_all_return())
        if all_return is None:
            return None
        whole_time = self.df.index.max() - self.df.index.min()
        days = whole_time.days
        if days <= 0:
            return None
        weeks = days / 7
        annual_return_ = (1 + all_return) ** (52 / weeks) - 1
        self.metrics['annual_return'] = annual_return_
        return annual_return_

    def calculate_annual_volatility(self):
        """计算年化波动率（倒序数据：最新日期在前）"""
        # 使用 dropna() 移除 NaN 值（最后一行会被移除，因为 shift(-1) 产生 NaN）
        weekly_returns = self.df["周收益率"].dropna()
        if len(weekly_returns) == 0:
            self.metrics['annual_volatility'] = None
            return None

        # 计算样本标准差
        variance = weekly_returns.var()  # pandas var() 使用 N-1
        standard = variance ** 0.5 # type: ignore

        # 年化波动率 = 周波动率 * sqrt(52)
        annual_volatility = standard * (52 ** 0.5)
        self.metrics['annual_volatility'] = annual_volatility
        return annual_volatility

    def calculate_sharpe_ratio(self):
        """计算夏普比率"""
        annual_return = self.metrics.get('annual_return')
        annual_volatility = self.metrics.get('annual_volatility')
        if annual_return is None or annual_volatility is None:
            return None
        sharpe_ratio = (annual_return - self.risk_free_rate) / annual_volatility
        self.metrics['sharpe_ratio'] = sharpe_ratio
        return sharpe_ratio

    def calculate_max_drawdown(self):
        """计算最大回撤（倒序数据：最新日期在前）"""
        # 先转为正序数据，以便正确计算 expanding().max()
        df_sorted = self.df.sort_index()
        df_sorted["历史最高值"] = df_sorted["单位净值"].expanding().max()
        df_sorted["每周回撤"] = df_sorted["单位净值"] / df_sorted["历史最高值"] - 1

        # 找到最小回撤（最大跌幅）
        max_drawback = df_sorted["每周回撤"].min()
        max_drawback_date = df_sorted["每周回撤"].idxmin()

        self.metrics['max_drawback'] = max_drawback
        self.metrics['max_drawback_date'] = max_drawback_date
        return max_drawback, max_drawback_date

    def calculate_1year_max_drawdown(self):
        """计算近一年最大回撤（倒序数据：最新日期在前）"""
        end_date = self.df.index.max()
        start_date = end_date - pd.Timedelta(days=365)

        # 获取近一年的数据
        df_1year = self.df[self.df.index >= start_date].copy()

        if len(df_1year) == 0:
            self.metrics['max_drawback_1year'] = None
            self.metrics['max_drawback_date_1year'] = None
            return None, None

        # 获取上一年最后一天的数据作为"历史最高"的起点
        # 这样才能正确计算从上一高点延续至今的回撤
        df_before = self.df[self.df.index < start_date]
        if len(df_before) > 0:
            # 上一年最后一天的净值作为初始历史最高
            prev_last_nav = df_before["单位净值"].iloc[-1]  # 倒序数据，最后一行是最早，所以iloc[-1]是最后一天
            # 将上一年最后一天的数据加入计算
            prev_day = pd.DataFrame({
                '单位净值': [prev_last_nav],
                '近一年历史最高': [prev_last_nav]  # 初始历史最高就是它自己
            }, index=[df_before.index[-1]])
            df_with_prev = pd.concat([prev_day, df_1year.sort_index()])
        else:
            df_with_prev = df_1year.sort_index()

        # 计算近一年内的历史最高值（从扩展数据中计算）
        df_with_prev["近一年历史最高"] = df_with_prev["单位净值"].expanding().max()

        # 计算回撤（只统计近一年内的）
        df_1year_sorted = df_1year.sort_index()
        for idx in df_1year_sorted.index:
            if idx in df_with_prev.index:
                df_1year_sorted.loc[idx, "近一年回撤"] = df_with_prev.loc[idx, "单位净值"] / df_with_prev.loc[idx, "近一年历史最高"] - 1 # type: ignore

        # 找到近一年内的最大回撤
        max_drawback_1year = df_1year_sorted["近一年回撤"].min()
        max_drawback_date_1year = df_1year_sorted["近一年回撤"].idxmin()

        self.metrics['max_drawback_1year'] = max_drawback_1year
        self.metrics['max_drawback_date_1year'] = max_drawback_date_1year
        return max_drawback_1year, max_drawback_date_1year

    def get_annual_returns(self):
        """计算年度收益率（倒序数据：最新日期在前）"""
        records = []
        # 获取所有年份，按时间正序排列
        years = sorted([int(y) for y in self.df.index.year.unique() if pd.notna(y)])  # type: ignore

        # 初始为最早年份的年初价格（DataFrame最后一行的净值，即最早日期）
        last_year_end = self.df["单位净值"].iloc[-1]

        # 按时间正序遍历年份
        for year in years:
            year_data = self.df[self.df.index.year == year]  # type: ignore
            start_price = last_year_end  # 年初价格 = 上一年的年末价格
            end_price = year_data['单位净值'].iloc[0]  # 年末价格 = 该年最新净值（第一行）
            total_return = (end_price / start_price) - 1
            records.append({
                'year': int(year),
                'start_price': round(start_price, 4),
                'end_price': round(end_price, 4),
                'annual_return': f"{total_return:.2%}"
            })
            last_year_end = end_price  # 更新为今年的年末价格
        return pd.DataFrame(records).set_index('year')

    def get_annual_max_drawdown(self, monthly: bool = False):
        """
        计算年度最大回撤（倒序数据：最新日期在前）

        Args:
            monthly: 是否使用月度数据计算

        Returns:
            DataFrame: 年度最大回撤数据
        """
        records = []
        years = sorted([int(y) for y in self.df.index.year.unique() if pd.notna(y)])  # type: ignore

        for i, year in enumerate(years):
            if i == 0:
                # 第一年：取该年的所有数据
                g = self.df[self.df.index.year == year].copy()  # type: ignore
            else:
                # 合并上一年年末和当年所有数据（倒序数据需要这样处理）
                prev_year_data = self.df[self.df.index.year == years[i-1]]  # type: ignore
                current_year_data = self.df[self.df.index.year == year]  # type: ignore
                # 取上一年最后一条记录作为起始点
                prev_last = prev_year_data.iloc[-1:].copy()
                # 合并数据
                g = pd.concat([prev_last, current_year_data])

            # 反转为正序（升序），以便正确计算expanding max
            g = g.sort_index()

            if monthly:
                # 月度数据需要重采样
                g = g.resample('ME').last()

            g["年内最高"] = g["单位净值"].expanding().max()
            g['年内回撤'] = g['单位净值'] / g["年内最高"] - 1

            this_year_mask = (g.index.year == year)  # type: ignore
            this_year_data = g[this_year_mask]

            if this_year_data.empty:
                continue

            min_dd = this_year_data['年内回撤'].min()
            trough = this_year_data['年内回撤'].idxmin()
            peak = g.loc[:trough, '单位净值'].idxmax()

            records.append({
                'year': int(year),
                'max_drawdown': f"{abs(min_dd):.2%}",
                'peak_date': peak.strftime('%Y-%m-%d'),
                'trough_date': trough.strftime('%Y-%m-%d')
            })

        return pd.DataFrame(records).set_index('year')

    def get_monthly_return_matrix(self):
        """计算成立以来月度收益矩阵（倒序数据：最新日期在前）"""
        records = []
        # 获取所有年月组合，按时间正序排列
        year_months = sorted([(int(y), int(m)) for y, m in self.df.groupby([self.df.index.year, self.df.index.month]).groups.keys() if pd.notna(y) and pd.notna(m)])  # type: ignore

        # 初始为最早月末价格（DataFrame最后一行的净值，即最早日期）
        last_month_end = self.df['单位净值'].iloc[-1]

        # 按时间正序遍历年月
        for (year, month) in year_months:
            month_data = self.df[(self.df.index.year == year) & (self.df.index.month == month)]  # type: ignore
            start_price = last_month_end  # 月初价格 = 上月末价格
            end_price = month_data['单位净值'].iloc[0]  # 月末价格 = 该月最新净值（第一行）
            month_return = (end_price / start_price) - 1
            records.append({
                'year': int(year),
                'month': month,
                'monthly_return': month_return
            })
            last_month_end = end_price  # 更新为本月末价格

        rec_df = pd.DataFrame(records)
        matrix = rec_df.pivot(index='year', columns='month', values='monthly_return').sort_index()
        matrix = matrix.reindex(columns=range(1, 13))
        annual_pct = rec_df.groupby('year')['monthly_return'].apply(lambda s: (s.add(1).prod() - 1) * 100)
        matrix = (matrix * 100).round(4)
        matrix['全年收益'] = annual_pct.reindex(matrix.index).round(4)
        return matrix

    def get_product_info(self):
        """获取产品信息（产品名称、最新净值日期、最新净值）"""
        if self.df.empty:
            return None, None, None

        product_name = list(self.products.values())[0]['name'] if self.products else None
        latest_nav_date = self.df.index.max()
        latest_nav = self.df['单位净值'].iloc[0]

        return product_name, latest_nav_date, latest_nav

    def build_metrics_df(self):
        """构建业绩指标DataFrame"""
        product_name, latest_nav_date, latest_nav = self.get_product_info()
        product_code = list(self.products.keys())[0] if self.products else None

        # 格式化日期为年月日
        def format_date(d):
            if d is None:
                return None
            if hasattr(d, 'strftime'):
                return d.strftime('%Y-%m-%d')
            return str(d)

        return pd.DataFrame([{
            '产品名称': product_name,
            '产品代码': product_code,
            '最新净值日期': format_date(latest_nav_date),
            '最新净值': latest_nav,
            '成立以来收益率': self.metrics.get('all_return'),
            '年化收益率': self.metrics.get('annual_return'),
            '年化波动率': self.metrics.get('annual_volatility'),
            '夏普比率': self.metrics.get('sharpe_ratio'),
            '最大回撤': self.metrics.get('max_drawback'),
            '最大回撤日期': format_date(self.metrics.get('max_drawback_date')),
            '近一年最大回撤': self.metrics.get('max_drawback_1year'),
            '近一年最大回撤日期': format_date(self.metrics.get('max_drawback_date_1year')),
        }])

    def run_all_calculations(self):
        """执行所有计算"""
        self.calculate_weekly_return()
        self.calculate_all_return()
        self.calculate_annual_return()
        self.calculate_annual_volatility()
        self.calculate_sharpe_ratio()
        self.calculate_max_drawdown()
        self.calculate_1year_max_drawdown()

    def save_to_excel(self, output_path: str):
        """
        保存单个产品的详细结果到Excel文件

        Args:
            output_path: 输出文件路径
        """
        metrics_df = self.build_metrics_df()
        annual_returns_df = self.get_annual_returns()
        weekly_dd = self.get_annual_max_drawdown(monthly=False)
        monthly_dd = self.get_annual_max_drawdown(monthly=True)
        monthly_matrix = self.get_monthly_return_matrix()

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            metrics_df.to_excel(writer, sheet_name='业绩指标计算', index=False)
            annual_returns_df.to_excel(writer, sheet_name='年度收益率')
            weekly_dd.to_excel(writer, sheet_name='周频计算历史最大回撤')
            monthly_dd.to_excel(writer, sheet_name='月频计算历史最大回撤')
            monthly_matrix.to_excel(writer, sheet_name='成立以来月度收益')

            wb = writer.book
            sheet_map = {
                '业绩指标计算': metrics_df,
                '年度收益率': annual_returns_df,
                '周频计算历史最大回撤': weekly_dd,
                '月频计算历史最大回撤': monthly_dd,
                '成立以来月度收益': monthly_matrix
            }

            for sheet_name, df_tmp in sheet_map.items():
                if df_tmp is None or df_tmp.empty:
                    continue
                ws = writer.sheets.get(sheet_name)
                if ws is None:
                    ws = wb[sheet_name]
                tmp = df_tmp.reset_index()
                for i, col in enumerate(tmp.columns, 1):
                    try:
                        # 计算该列数据的最大长度
                        max_data_len = tmp[col].astype(str).map(len).max()
                    except Exception:
                        max_data_len = 0
                    header_len = len(str(col))
                    # 列宽 = max(数据最大长度, 表头长度) + 额外缓冲
                    adjusted_width = max(max_data_len, header_len, 8) + 4
                    col_letter = get_column_letter(i)
                    ws.column_dimensions[col_letter].width = min(adjusted_width, 50)  # 最大50防止过长

                # 设置数值格式（year列保持整数格式）
                try:
                    year_col_idx = None
                    # 找到year列的位置
                    for col_idx, col_name in enumerate(tmp.columns, 1):
                        if str(col_name).lower() == 'year':
                            year_col_idx = col_idx
                            break
                    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in r:
                            if isinstance(cell.value, (int, float)):
                                # year列保持整数格式
                                if year_col_idx and cell.column == year_col_idx:
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.0000'
                except Exception:
                    pass

        print(f'已保存文件: {output_path}')

    def print_summary(self):
        """打印计算结果摘要"""
        product_name, latest_nav_date, latest_nav = self.get_product_info()
        # 格式化日期为年月日
        def format_date(d):
            if d is None:
                return 'N/A'
            if hasattr(d, 'strftime'):
                return d.strftime('%Y-%m-%d')
            return str(d)

        print(f"\n产品名称: {product_name}")
        print(f"产品代码: {list(self.products.keys())[0] if self.products else 'N/A'}")
        print(f"最新净值日期: {format_date(latest_nav_date)}")
        print(f"最新净值: {latest_nav:.4f}")

        all_return = self.metrics.get('all_return')
        annual_return = self.metrics.get('annual_return')
        annual_volatility = self.metrics.get('annual_volatility')
        sharpe_ratio = self.metrics.get('sharpe_ratio')
        max_drawback = self.metrics.get('max_drawback')
        max_drawback_date = self.metrics.get('max_drawback_date')
        max_drawback_1year = self.metrics.get('max_drawback_1year')
        max_drawback_date_1year = self.metrics.get('max_drawback_date_1year')

        print()
        if all_return is not None:
            print(f"成立以来收益率：{all_return:.2%}")
        if annual_return is not None:
            print(f"年化收益率是 {annual_return:.2%}")
        if annual_volatility is not None:
            print(f"年化波动率为{annual_volatility:.2%}")
        if sharpe_ratio is not None:
            print(f"夏普比率是{sharpe_ratio:.4}")
        if max_drawback is not None and max_drawback_date is not None:
            print(f"最大回撤发生在{format_date(max_drawback_date)}，为{-max_drawback:.2%}")
        if max_drawback_1year is not None and max_drawback_date_1year is not None:
            print(f"近一年最大回撤发生在{format_date(max_drawback_date_1year)}，为{-max_drawback_1year:.2%}")
